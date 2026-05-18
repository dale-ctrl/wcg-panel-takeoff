"""
WCG Panel Takeoff Tool v2
Extracts panel schedules from SolidWorks STEP files and exports to Smart Space supplier template.

v2 improvements:
- Orientation-based classification (uses XYZ axis, not just sorted dimensions)
- Position-based top/base/back/door distinction
- LH vs RH side panel detection
- Per-cluster classification for mixed cabinet assemblies
- Non-rectangular body detection (volume ratio check)
- Face-count filtering for cylindrical/complex hardware
- Material assignment per part type
- Optional edgebanding deductions
"""

import streamlit as st
import cadquery as cq
import pandas as pd
import numpy as np
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import Counter
from datetime import date
import tempfile
import os
import io
import re

st.set_page_config(page_title="WCG Panel Takeoff", layout="wide", page_icon="\U0001fab5")

PART_TYPES = [
    "Side Panel LH", "Side Panel RH", "Back", "Top", "Base",
    "Shelf", "Door", "Sliding Door", "End Panel", "Top Rail",
    "Plinth", "Top Pelmet", "Base Panel", "Infill Rail", "Infill Rails",
    "Worktop", "Upstand", "Other"
]

MATERIAL_PRESETS = {
    "18mm Sorano Oak H1334": "18mm Sorano Oak H1334",
    "25mm Sorano Oak H1334": "25mm Sorano Oak H1334",
    "18mm MFC White": "18mm MFC White",
    "25mm MFC White": "25mm MFC White",
    "18mm MFMDF": "18mm MFMDF",
    "25mm MFMDF": "25mm MFMDF",
    "38mm Egger Worktops Oak Butcherblock": "38mm Egger Worktops Oak Butcherblock",
    "18mm Egger Upstand Oak Butcherblock": "18mm Egger Upstand Oak Butcherblock",
    "Custom...": "",
}
MATERIAL_LIST = list(MATERIAL_PRESETS.keys())

EDGING_CODES = ["LLWW", "L", "LW", "LLW", "W", "0"]
EDGING_TYPES = [
    "2MM-STANDARD-ABS-MATCHING",
    "1MM-STANDARD-ABS-MATCHING",
    "0.4MM-STANDARD-ABS-MATCHING",
    "None",
]
EDGING_PROFILES = ["2mm radius", "Square", "1mm radius"]

MIN_PANEL_SIZE_MM = 150
MAX_PANEL_FACES = 30
VOLUME_RATIO_WARN = 0.85


def extract_bodies_from_step(file_bytes, filename):
    with tempfile.NamedTemporaryFile(suffix=".STEP", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        result = cq.importers.importStep(tmp_path)
        solids = result.solids().vals()
    finally:
        os.unlink(tmp_path)

    assy_name = os.path.splitext(filename)[0]
    for suffix in ["_Assy", " Assy", "_Assembly", " Assembly"]:
        assy_name = assy_name.replace(suffix, "")
    assy_name = assy_name.replace("_", " ")

    bodies = []
    hardware_count = 0
    for solid in solids:
        bb = solid.BoundingBox()
        xlen = round(bb.xlen, 1)
        ylen = round(bb.ylen, 1)
        zlen = round(bb.zlen, 1)
        dims_sorted = sorted([xlen, ylen, zlen], reverse=True)
        face_count = len(solid.Faces())
        bb_vol = xlen * ylen * zlen
        actual_vol = solid.Volume()
        vol_ratio = round(actual_vol / bb_vol, 3) if bb_vol > 0 else 0

        if dims_sorted[0] < MIN_PANEL_SIZE_MM:
            hardware_count += 1
            continue
        if face_count > MAX_PANEL_FACES:
            hardware_count += 1
            continue

        min_dim = min(xlen, ylen, zlen)
        if min_dim == xlen:
            thin_axis = "X"
            length = max(ylen, zlen)
            width = min(ylen, zlen)
        elif min_dim == ylen:
            thin_axis = "Y"
            length = max(xlen, zlen)
            width = min(xlen, zlen)
        else:
            thin_axis = "Z"
            length = max(xlen, ylen)
            width = min(xlen, ylen)

        bodies.append({
            "xlen": xlen, "ylen": ylen, "zlen": zlen,
            "length": length, "width": width, "thickness": min_dim,
            "thin_axis": thin_axis,
            "x_ctr": round((bb.xmin + bb.xmax) / 2, 1),
            "y_ctr": round((bb.ymin + bb.ymax) / 2, 1),
            "z_ctr": round((bb.zmin + bb.zmax) / 2, 1),
            "face_count": face_count, "vol_ratio": vol_ratio,
            "source_file": filename, "assembly": assy_name,
        })
    return bodies, hardware_count, assy_name


def detect_axis_convention(bodies):
    """Auto-detect CAD axis convention (works for SolidWorks, VectorWorks, etc).
    Returns: {'sides': axis, 'vertical': axis, 'depth': axis}
    where axis is 'X', 'Y', or 'Z'."""
    if not bodies:
        return {"sides": "X", "vertical": "Y", "depth": "Z"}

    # Height axis: the axis that tall panels' longest dimension runs along.
    # Only use genuinely tall panels (>1500mm) to avoid shelves skewing the vote.
    tall = [b for b in bodies if b["length"] > 1500]
    if not tall:
        tall = [b for b in bodies if b["length"] > 800]
    if not tall:
        return {"sides": "X", "vertical": "Y", "depth": "Z"}

    votes = {"X": 0, "Y": 0, "Z": 0}
    for b in tall:
        thin = b["thin_axis"]
        if thin == "X":
            votes["Y" if b["ylen"] >= b["zlen"] else "Z"] += 1
        elif thin == "Y":
            votes["X" if b["xlen"] >= b["zlen"] else "Z"] += 1
        else:
            votes["X" if b["xlen"] >= b["ylen"] else "Y"] += 1

    height_axis = max(votes, key=votes.get)
    remaining = [a for a in "XYZ" if a != height_axis]
    ax_a, ax_b = remaining

    # Depth axis has the largest-area thin panel (the back).
    thin_a = [b for b in bodies if b["thin_axis"] == ax_a and b["width"] > 130]
    thin_b = [b for b in bodies if b["thin_axis"] == ax_b and b["width"] > 130]
    def _max_area(panels):
        return max((b["length"] * b["width"] for b in panels), default=0)

    if _max_area(thin_a) >= _max_area(thin_b):
        return {"sides": ax_b, "vertical": height_axis, "depth": ax_a}
    else:
        return {"sides": ax_a, "vertical": height_axis, "depth": ax_b}


def _get_role(thin_axis, convention):
    """Map a body's thin axis to its functional role: sides/vertical/depth."""
    for role, axis in convention.items():
        if axis == thin_axis:
            return role
    return "unknown"


def _pos(body, axis):
    """Get centre position of body along a given axis."""
    return body[f"{axis.lower()}_ctr"]


def classify_body(body, all_bodies, convention):
    """Classify a panel using orientation role and 3D position.
    Works with any CAD axis convention via the detected convention mapping.

    Roles:
        sides    (thin faces left/right)  -> Side Panels, End Panels
        vertical (thin faces up/down)     -> Tops, Bases, Shelves, Rails
        depth    (thin faces front/back)  -> Backs, Doors, Plinths
    """
    role = _get_role(body["thin_axis"], convention)
    thin_ax = body["thin_axis"]
    length = body["length"]
    width = body["width"]
    thickness = body["thickness"]
    ratio = length / width if width > 0 else 999
    notes = []

    if body["vol_ratio"] < VOLUME_RATIO_WARN:
        notes.append(f"Non-rectangular (vol {body['vol_ratio']:.0%}) -- check shape")

    # Worktops
    if thickness > 30:
        return "Worktop", "0", notes

    # Very narrow = rails/plinths
    if width <= 130:
        if length > 2500:
            notes.append("Full-width -- may need splitting per cabinet")
        if role == "depth":
            return "Plinth", "LLWW", notes
        if role == "vertical":
            return ("Top Rail", "L", notes) if thickness >= 20 else ("Plinth", "LLWW", notes)
        return "Top Rail", "L", notes

    # Full-width spanning panels
    same_role = [b for b in all_bodies if _get_role(b["thin_axis"], convention) == role and b["width"] > 130]
    same_lens = [b["length"] for b in same_role]
    if length > 2500 and any(l < length * 0.7 for l in same_lens):
        notes.append("Full-width -- may need splitting per cabinet")
        if role == "vertical":
            return "Top Rail", "L", notes
        return "End Panel", "LLWW", notes

    # === SIDES ROLE: Side / End panels ===
    if role == "sides":
        side_panels = [b for b in all_bodies if _get_role(b["thin_axis"], convention) == "sides" and b["width"] > 130]
        if side_panels:
            all_pos = [_pos(b, thin_ax) for b in side_panels]
            pos_min, pos_max = min(all_pos), max(all_pos)
            heights = sorted(set(b["length"] for b in side_panels), reverse=True)
            this_pos = _pos(body, thin_ax)
            at_extreme = abs(this_pos - pos_min) < 50 or abs(this_pos - pos_max) < 50

            if len(heights) >= 2 and at_extreme and length >= heights[0] - 5:
                return "End Panel", "LLWW", notes
            if at_extreme and len(heights) == 1 and thickness >= 20:
                return "End Panel", "LLWW", notes

            # LH vs RH
            non_end = [b for b in side_panels
                       if not (len(heights) >= 2 and b["length"] >= heights[0] - 5
                               and (abs(_pos(b, thin_ax) - pos_min) < 50 or abs(_pos(b, thin_ax) - pos_max) < 50))]
            if non_end:
                other_pos = [_pos(b, thin_ax) for b in non_end if abs(_pos(b, thin_ax) - this_pos) > 10]
                if other_pos:
                    nearest = min(other_pos, key=lambda p: abs(p - this_pos))
                    if abs(this_pos - nearest) < 200:
                        return ("Side Panel LH", "L", notes) if this_pos < nearest else ("Side Panel RH", "L", notes)

        return "Side Panel LH", "L", notes

    # === VERTICAL ROLE: Tops, Bases, Shelves ===
    if role == "vertical":
        vert_panels = [b for b in all_bodies if _get_role(b["thin_axis"], convention) == "vertical" and b["width"] > 130]
        if vert_panels:
            all_pos = [_pos(b, thin_ax) for b in vert_panels]
            pos_max, pos_min = max(all_pos), min(all_pos)
            pos_range = pos_max - pos_min
            this_pos = _pos(body, thin_ax)

            if thickness >= 20:
                return "Shelf", "L", notes
            if pos_range > 50 and abs(this_pos - pos_max) < 50:
                return "Top", "L", notes
            if pos_range > 50 and abs(this_pos - pos_min) < 50:
                return "Base", "L", notes
            if pos_range > 50:
                return "Shelf", "L", notes
        return "Top", "L", notes

    # === DEPTH ROLE: Backs, Doors ===
    if role == "depth":
        depth_panels = [b for b in all_bodies if _get_role(b["thin_axis"], convention) == "depth" and b["width"] > 130]
        if depth_panels:
            all_pos = [_pos(b, thin_ax) for b in depth_panels]
            pos_min, pos_max = min(all_pos), max(all_pos)
            pos_range = pos_max - pos_min
            this_pos = _pos(body, thin_ax)

            # Back = rearmost. Convention: back is at max depth for SolidWorks (high X in VW).
            # We determine "rear" as the position where the largest-area panel sits.
            areas_by_pos = {}
            for b in depth_panels:
                p = _pos(b, thin_ax)
                a = b["length"] * b["width"]
                areas_by_pos[p] = max(areas_by_pos.get(p, 0), a)
            back_pos = max(areas_by_pos, key=areas_by_pos.get)

            if pos_range > 50 and abs(this_pos - back_pos) < 50:
                return "Back", "0", notes
            if pos_range > 50:
                return "Door", "LLWW", notes
            # Single depth position — use area
            max_area = max(b["length"] * b["width"] for b in depth_panels)
            if length * width >= max_area * 0.9 and ratio < 2.0:
                return "Back", "0", notes
            return "Door", "LLWW", notes
        return "Back", "0", notes

    return "Other", "L", notes


# --- PDF parsing (WCG supplier drawings) ---------------------------------

PDF_DIM_UPRIGHT = re.compile(r"^\d{2,4}\.\d$")
PDF_DIM_REVERSED = re.compile(r"^\d\.\d{2,4}$")
PDF_MIN_DIM = 50
PDF_DESC_BLOCK_W = 150
PDF_DESC_BLOCK_H = 130

PDF_FIELD_RE = {
    "material": re.compile(r"^Material:\s*(.*)$", re.I),
    "thickness": re.compile(r"^Thickness:\s*(.*)$", re.I),
    "edge_finish": re.compile(r"^Edge\s*Finish:\s*(.*)$", re.I),
    "qty": re.compile(r"^Qty:\s*(.*)$", re.I),
    "notes": re.compile(r"^Notes:\s*(.*)$", re.I),
}
PDF_DELIM_PATS = list(PDF_FIELD_RE.values()) + [
    re.compile(r"^Colour:", re.I),
    re.compile(r"^Cutting:", re.I),
    re.compile(r"^Config:", re.I),
]

PDF_TEMPLATE_PATS = [
    re.compile(r"drill\s+template", re.I),
    re.compile(r"\btemplate\b", re.I),
]

PDF_NAME_TO_PART_TYPE = [
    (re.compile(r"worktop", re.I), "Worktop"),
    (re.compile(r"upstand", re.I), "Upstand"),
    (re.compile(r"pelmet", re.I), "Top Pelmet"),
    (re.compile(r"plinth", re.I), "Plinth"),
    (re.compile(r"back\s*panel|\bback\b", re.I), "Back"),
    (re.compile(r"lh\s*panel|\blh\b|left\s*panel", re.I), "Side Panel LH"),
    (re.compile(r"rh\s*panel|\brh\b|right\s*panel", re.I), "Side Panel RH"),
    (re.compile(r"end\s*panel|\bend\b", re.I), "End Panel"),
    (re.compile(r"sliding\s*door", re.I), "Sliding Door"),
    (re.compile(r"\bdoor\b", re.I), "Door"),
    (re.compile(r"shelf", re.I), "Shelf"),
    (re.compile(r"upright|divider", re.I), "Side Panel LH"),
    (re.compile(r"base\s*panel|\bbase\b", re.I), "Base"),
    (re.compile(r"top\s*panel|\btop\b", re.I), "Top"),
    (re.compile(r"\brail\b", re.I), "Top Rail"),
    (re.compile(r"spacer", re.I), "Other"),
]


def classify_pdf_name(name):
    for pat, pt in PDF_NAME_TO_PART_TYPE:
        if pat.search(name):
            return pt
    return "Other"


def pdf_edge_finish_to_code(text):
    t = (text or "").lower()
    if "all around" in t:
        return "LLWW"
    if "no edging" in t or t.strip() == "raw":
        return "0"
    if "where indicated" in t:
        return "L"
    return "L"


def _pdf_parse_dim(text, upright):
    text = text.strip()
    if PDF_DIM_UPRIGHT.fullmatch(text):
        try:
            v = float(text)
            return v if v >= PDF_MIN_DIM else None
        except ValueError:
            return None
    if not upright and PDF_DIM_REVERSED.fullmatch(text):
        try:
            v = float(text[::-1])
            return v if v >= PDF_MIN_DIM else None
        except ValueError:
            return None
    return None


def _pdf_parse_block(text):
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    result = {"name": "", "material": "", "thickness": 0.0,
              "edge_finish": "", "qty": 1, "raw_qty": "", "notes": ""}
    start = -1
    for i, line in enumerate(lines):
        if line.startswith("Component:"):
            start = i
            break
    if start < 0:
        return result

    field_starts = [j for j in range(start + 1, len(lines))
                    if any(p.match(lines[j]) for p in PDF_DELIM_PATS)]

    if field_starts:
        name_lines = lines[start + 1:field_starts[0]]
    else:
        name_lines = lines[start + 1:]
    name_lines = [l for l in name_lines if "$PRPVIEW" not in l]
    result["name"] = " ".join(name_lines).strip()

    for idx, sj in enumerate(field_starts):
        line = lines[sj]
        end_j = field_starts[idx + 1] if idx + 1 < len(field_starts) else len(lines)
        matched_key = None
        first_val = ""
        for key, pat in PDF_FIELD_RE.items():
            m = pat.match(line)
            if m:
                matched_key = key
                first_val = m.group(1).strip()
                break
        if matched_key is None:
            continue
        parts = []
        if first_val and "$PRPVIEW" not in first_val:
            parts.append(first_val)
        for k in range(sj + 1, end_j):
            v = lines[k].strip()
            if v and "$PRPVIEW" not in v:
                parts.append(v)
        value = " ".join(parts).strip()
        if matched_key == "thickness":
            tm = re.search(r"(\d+\.?\d*)", value)
            result["thickness"] = float(tm.group(1)) if tm else 0.0
        elif matched_key == "qty":
            result["raw_qty"] = value
            qm = re.search(r"(\d+)", value)
            result["qty"] = int(qm.group(1)) if qm else 1
        else:
            result[matched_key] = value
    return result


def _pdf_crop_block(page, anchor):
    bbox = (
        max(0, anchor["x"] - 3),
        max(0, anchor["y"] - 3),
        min(page.width, anchor["x"] + PDF_DESC_BLOCK_W),
        min(page.height, anchor["y"] + PDF_DESC_BLOCK_H + 30),
    )
    try:
        return page.crop(bbox).extract_text() or ""
    except Exception:
        return ""


def _pdf_compute_sections(anchors, page_width, page_height, row_tol=100):
    if not anchors:
        return []
    sorted_anchors = sorted(enumerate(anchors), key=lambda t: t[1]["y"])
    rows = []
    for idx, a in sorted_anchors:
        if rows and abs(a["y"] - rows[-1][-1][1]["y"]) < row_tol:
            rows[-1].append((idx, a))
        else:
            rows.append([(idx, a)])
    row_centers = [sum(t[1]["y"] for t in r) / len(r) for r in rows]
    sections = [None] * len(anchors)
    for ri, row in enumerate(rows):
        row.sort(key=lambda t: t[1]["x"])
        y_top = 0 if ri == 0 else (row_centers[ri - 1] + row_centers[ri]) / 2
        y_bot = page_height if ri == len(rows) - 1 else (row_centers[ri] + row_centers[ri + 1]) / 2
        for i, (idx, a) in enumerate(row):
            x_left = 0 if i == 0 else a["x"] - 5
            x_right = page_width if i + 1 == len(row) else row[i + 1][1]["x"]
            sections[idx] = (x_left, y_top, x_right, y_bot)
    return sections


def _pdf_assembly_name(pdf, fallback):
    for page in pdf.pages:
        text = page.extract_text() or ""
        for line in text.split("\n"):
            m = re.search(r"Description:\s*([^\r\n]+)", line)
            if not m:
                continue
            name = m.group(1).strip()
            if not name or "$PRPVIEW" in name:
                continue
            name = re.sub(r"\s+Product no.*$", "", name, flags=re.I).strip()
            name = re.sub(r"\s+Job no.*$", "", name, flags=re.I).strip()
            if name:
                return name
    return fallback


def extract_panels_from_pdf(file_bytes, filename):
    """Parse a WCG supplier-drawing PDF and return (bodies, assembly_name).

    Each body has: name, material, thickness, edge_finish, qty, raw_qty,
    length, width, page, source_file.
    """
    bodies = []
    assy_name = os.path.splitext(filename)[0]
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        assy_name = _pdf_assembly_name(pdf, assy_name)
        for page_idx, page in enumerate(pdf.pages):
            words = page.extract_words(
                use_text_flow=False, keep_blank_chars=False,
                x_tolerance=2, y_tolerance=2, extra_attrs=["upright"],
            )
            raw_anchors = [
                {"x": w["x0"], "y": w["top"]}
                for w in words
                if w["text"] == "Component:" and w["x0"] < page.width - 80
            ]
            if not raw_anchors:
                continue

            parsed = []
            for a in raw_anchors:
                block_text = _pdf_crop_block(page, a)
                comp = _pdf_parse_block(block_text)
                if not comp["name"] or "$PRPVIEW" in comp["name"]:
                    continue
                if any(p.search(comp["name"]) for p in PDF_TEMPLATE_PATS):
                    continue
                parsed.append((a, comp))
            if not parsed:
                continue

            anchors = [p[0] for p in parsed]
            block_bboxes = [
                (a["x"] - 3, a["y"] - 3,
                 a["x"] + PDF_DESC_BLOCK_W, a["y"] + PDF_DESC_BLOCK_H)
                for a in anchors
            ]
            dim_words = []
            for w in words:
                val = _pdf_parse_dim(w["text"], w.get("upright", True))
                if val is None:
                    continue
                cx = (w["x0"] + w["x1"]) / 2
                cy = (w["top"] + w["bottom"]) / 2
                if any(bb[0] <= cx <= bb[2] and bb[1] <= cy <= bb[3] for bb in block_bboxes):
                    continue
                if cx > page.width * 0.5 and cy > page.height * 0.85:
                    continue
                dim_words.append({"value": val, "x": cx, "y": cy})

            sections = _pdf_compute_sections(anchors, page.width, page.height)
            section_dims = [[] for _ in anchors]
            for d in dim_words:
                for i, sec in enumerate(sections):
                    if sec and sec[0] <= d["x"] <= sec[2] and sec[1] <= d["y"] <= sec[3]:
                        section_dims[i].append(d["value"])
                        break

            for i, (anchor, comp) in enumerate(parsed):
                uniq = sorted(set(section_dims[i]), reverse=True)
                length = uniq[0] if uniq else 0
                width = uniq[1] if len(uniq) > 1 else 0
                bodies.append({
                    "name": comp["name"],
                    "material": comp["material"],
                    "thickness": comp["thickness"],
                    "edge_finish": comp["edge_finish"],
                    "qty": comp["qty"],
                    "raw_qty": comp["raw_qty"],
                    "length": length,
                    "width": width,
                    "page": page_idx + 1,
                    "source_file": filename,
                })
    return bodies, assy_name


def process_pdf_assembly(bodies, assy_name):
    """Convert PDF bodies to row dicts in the same shape as process_assembly."""
    rows = []
    for b in bodies:
        part_type = classify_pdf_name(b["name"])
        edging = pdf_edge_finish_to_code(b["edge_finish"])
        notes = []
        if not b["length"] or not b["width"]:
            notes.append("Dimensions incomplete -- verify against drawing")
        if b.get("raw_qty") and re.search(r"\b(RH|LH|mirror)", b["raw_qty"], re.I):
            notes.append(f"Qty text: '{b['raw_qty']}' -- check for mirrored variant")
        rows.append({
            "part_type": part_type,
            "component_name": b["name"],
            "length": b["length"],
            "width": b["width"],
            "thickness": b["thickness"],
            "qty": b["qty"],
            "edging_code": edging,
            "edge_finish_text": b["edge_finish"],
            "auto_notes": "; ".join(notes),
            "page": b["page"],
        })
    return {assy_name: rows}


# --- End PDF parsing -----------------------------------------------------


def process_assembly(bodies, assy_name):
    if not bodies:
        return {}

    classified = []
    convention = detect_axis_convention(bodies)
    for body in bodies:
        part_type, edging, auto_notes = classify_body(body, bodies, convention)
        classified.append({**body, "part_type": part_type, "edging_code": edging, "auto_notes": auto_notes})

    dedup_key = lambda p: (p["part_type"], p["length"], p["width"], p["thickness"])
    counter = Counter()
    reps = {}
    notes_coll = {}
    for p in classified:
        key = dedup_key(p)
        counter[key] += 1
        if key not in reps:
            reps[key] = p
            notes_coll[key] = list(p["auto_notes"])
        else:
            for n in p["auto_notes"]:
                if n not in notes_coll[key]:
                    notes_coll[key].append(n)

    rows = []
    for key, qty in sorted(counter.items(),
                           key=lambda x: (PART_TYPES.index(x[0][0]) if x[0][0] in PART_TYPES else 99, -x[0][1])):
        p = reps[key]
        nts = list(notes_coll[key])
        rows.append({
            "part_type": key[0], "length": key[1], "width": key[2],
            "thickness": key[3], "qty": qty, "edging_code": p["edging_code"],
            "auto_notes": "; ".join(nts), "thin_axis": p["thin_axis"],
            "vol_ratio": p["vol_ratio"], "face_count": p["face_count"],
        })
    return {assy_name: rows}


def apply_edgeband_deductions(rows, deduction_mm, apply_deductions):
    if not apply_deductions or deduction_mm <= 0:
        return rows
    adjusted = []
    for row in rows:
        r = dict(row)
        code = r.get("edging_code", "0")
        d = deduction_mm
        if code == "LLWW":
            r["length"] -= 2 * d
            r["width"] -= 2 * d
        elif code == "L":
            r["length"] -= 2 * d
        elif code == "LW":
            r["length"] -= d
            r["width"] -= d
        elif code == "LLW":
            r["length"] -= 2 * d
            r["width"] -= d
        elif code == "W":
            r["width"] -= 2 * d
        adjusted.append(r)
    return adjusted


def build_smart_space_xlsx(project_settings, panel_groups):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    fnt_label = Font(name="Franklin Gothic Book", size=10.5)
    fnt_value = Font(name="Franklin Gothic Book", size=11)
    fnt_header = Font(name="Franklin Gothic Book", size=10.5, bold=True)
    fnt_section = Font(name="Franklin Gothic Book", size=16, bold=True)
    fnt_note = Font(name="Franklin Gothic Book", size=10)
    fnt_data = Font(name="Franklin Gothic Book", size=11)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    for col, w in {"A":41.3,"B":42.4,"C":14.1,"D":14.7,"E":13.0,"F":7.3,
                   "G":23.6,"H":32.6,"I":25.4,"J":22.0,"K":28.1}.items():
        ws.column_dimensions[col].width = w

    ws["J2"] = "**EDGING DIAGRAM"; ws["J2"].font = fnt_label
    ws["A3"] = "Customer Name:"; ws["A3"].font = fnt_label
    ws["B3"] = project_settings.get("customer_name", ""); ws["B3"].font = fnt_value; ws.merge_cells("B3:C3")
    ws["A4"] = "E-mail address:"; ws["A4"].font = fnt_label
    ws["B4"] = project_settings.get("email", ""); ws["B4"].font = fnt_value; ws.merge_cells("B4:C4")
    ws["D4"] = "Address:"; ws["D4"].font = fnt_label
    ws["E4"] = project_settings.get("address", ""); ws["E4"].font = fnt_value
    ws["A5"] = "Contact No.:"; ws["A5"].font = fnt_label
    ws["B5"] = project_settings.get("contact_no", ""); ws["B5"].font = fnt_value; ws.merge_cells("B5:C5")
    ws["D5"] = "Offcuts Required (Y/N):"; ws["D5"].font = fnt_label
    ws["F5"] = "Delivery (Y/N):"; ws["F5"].font = fnt_label; ws.merge_cells("F5:G5")
    ws["H5"] = project_settings.get("delivery_yn", "Y"); ws["H5"].font = fnt_value
    ws["A6"] = "Project Name:"; ws["A6"].font = fnt_label
    ws["B6"] = project_settings.get("project_name", ""); ws["B6"].font = fnt_value; ws.merge_cells("B6:C6")
    ws["D6"] = "Enquiry Date:"; ws["D6"].font = fnt_label
    ws["E6"] = date.today(); ws["E6"].font = fnt_value; ws["E6"].number_format = "DD/MM/YYYY"
    ws["F6"] = "Delivery street name & postcode:"; ws["F6"].font = fnt_label; ws.merge_cells("F6:G6")
    ws["H6"] = project_settings.get("delivery_postcode", ""); ws["H6"].font = fnt_value

    ws["D8"] = '*Finished size (mm) including edgebanding                                         *Grain runs with length dimension'
    ws["D8"].font = fnt_note; ws.merge_cells("D8:E8")
    ws["G8"] = '**Please see Edging diagram: "L" relates to Length, "W" to Width. Use number 0 for NONE.'
    ws["G8"].font = fnt_label
    ws["H8"] = '*** Use dropdown menus for Edgebanding Type and Edgebanding Profile'
    ws["H8"].font = fnt_note; ws.merge_cells("H8:I8")

    for col, text in [("A","PART DESCRIPTION (to match DXF)"),("B","SHEET MATERIAL DESCRIPTION"),
        ("C","CABINET No."),("D","LENGTH"),("E","WIDTH"),("F","QTY"),("G","EDGEBAND - LLWW"),
        ("H","EDGEBANDING TYPE"),("I","EDGEBANDING PROFILE"),("J","Notes"),("K","DXF Path")]:
        c = ws[f"{col}10"]; c.value = text; c.font = fnt_header; c.fill = header_fill
        c.border = thin_border; c.alignment = Alignment(horizontal="center", wrap_text=True)

    for col, text in {"A":"for reference (e.g Top)","B":" (include thickness)","C":"(if applicable)",
        "D":"(mm)*","E":"(mm)*","G":"Follow code LLWW**","I":"Square/Radius",
        "J":"(if applicable)","K":"Filled for Perfect Panels only"}.items():
        c = ws[f"{col}11"]; c.value = text; c.font = fnt_label; c.alignment = Alignment(horizontal="center")

    row = 12
    for group_name, group_panels in panel_groups.items():
        ws[f"A{row}"] = group_name; ws[f"A{row}"].font = fnt_section; row += 1
        for panel in group_panels:
            ws[f"A{row}"] = panel.get("part_type", ""); ws[f"A{row}"].font = fnt_data
            ws[f"B{row}"] = panel.get("material", ""); ws[f"B{row}"].font = fnt_data
            ws[f"C{row}"] = panel.get("cabinet_no", ""); ws[f"C{row}"].font = fnt_data
            ws[f"D{row}"] = int(panel.get("length", 0)); ws[f"D{row}"].font = fnt_data; ws[f"D{row}"].number_format = "0"
            ws[f"E{row}"] = int(panel.get("width", 0)); ws[f"E{row}"].font = fnt_data; ws[f"E{row}"].number_format = "0"
            ws[f"F{row}"] = int(panel.get("qty", 1)); ws[f"F{row}"].font = fnt_data
            ws[f"G{row}"] = panel.get("edging_code", "0"); ws[f"G{row}"].font = fnt_data
            if panel.get("edging_code", "0") != "0":
                ws[f"H{row}"] = panel.get("edging_type", ""); ws[f"H{row}"].font = fnt_data
                ws[f"I{row}"] = panel.get("edging_profile", ""); ws[f"I{row}"].font = fnt_data
            ws[f"J{row}"] = panel.get("notes", ""); ws[f"J{row}"].font = fnt_data
            for col in "ABCDEFGHIJK":
                ws[f"{col}{row}"].border = thin_border
            row += 1
        row += 1

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def main():
    st.title("\U0001fab5 WCG Panel Takeoff")
    st.caption("Extract panel schedules from SolidWorks STEP files -> Smart Space supplier template")

    with st.sidebar:
        st.header("Project Settings")
        project_name = st.text_input("Project Name", value="", placeholder="e.g. 14998 - BPS")
        customer_name = st.text_input("Customer Name", value="Westcountry Group Ltd")
        email = st.text_input("Email", value="tom@westcountrygroup.com")
        contact_no = st.text_input("Contact No.", value="")
        delivery_postcode = st.text_input("Delivery Postcode", value="")
        delivery_yn = st.selectbox("Delivery (Y/N)", ["Y", "N"])

        st.divider()
        st.header("Material per Part Type")
        mat_choices = [m for m in MATERIAL_PRESETS.values() if m]
        mat_side = st.selectbox("Side Panels", mat_choices, index=0, key="ms")
        mat_end = st.selectbox("End Panels", mat_choices, index=1, key="me")
        mat_back = st.selectbox("Backs", mat_choices, index=0, key="mb")
        mat_top = st.selectbox("Tops & Bases", mat_choices, index=0, key="mt")
        mat_shelf = st.selectbox("Shelves", mat_choices, index=1, key="msh")
        mat_door = st.selectbox("Doors", mat_choices, index=0, key="md")
        mat_rail = st.selectbox("Rails & Plinths", mat_choices, index=0, key="mr")
        custom_mat = st.text_input("Custom material (Other)", value="", key="mc")

        material_map = {
            "Side Panel LH": mat_side, "Side Panel RH": mat_side,
            "End Panel": mat_end, "Back": mat_back,
            "Top": mat_top, "Base": mat_top, "Shelf": mat_shelf,
            "Door": mat_door, "Sliding Door": mat_door,
            "Top Rail": mat_rail, "Plinth": mat_rail,
            "Top Pelmet": mat_rail, "Base Panel": mat_top,
            "Infill Rail": mat_rail, "Infill Rails": mat_rail,
            "Worktop": "38mm Egger Worktops Oak Butcherblock",
            "Upstand": "18mm Egger Upstand Oak Butcherblock",
            "Other": custom_mat or mat_side,
        }

        st.divider()
        st.header("Edgebanding Deductions")
        apply_ded = st.toggle("Apply edgeband deductions", value=False)
        ded_mm = 0.0
        if apply_ded:
            ded_mm = st.number_input("Deduction per edge (mm)", 0.0, 5.0, 2.0, 0.1)
            st.info(f"L = -{ded_mm*2:.1f}mm length | LLWW = -{ded_mm*2:.1f}mm both")

        default_etype = st.selectbox("Default edging type", EDGING_TYPES, index=0)
        default_eprof = st.selectbox("Default edging profile", EDGING_PROFILES, index=0)

    project_settings = {
        "project_name": project_name, "customer_name": customer_name,
        "email": email, "contact_no": contact_no,
        "delivery_postcode": delivery_postcode, "delivery_yn": delivery_yn, "address": "",
    }

    tab_step, tab_pdf, tab_manual = st.tabs(["STEP File Upload", "PDF Upload", "Manual Entry"])

    with tab_step:
        uploaded_files = st.file_uploader("Upload STEP files", type=["step","stp"], accept_multiple_files=True)
        if uploaded_files and st.button("Extract Panels", type="primary"):
            all_groups = {}; total_hw = 0
            progress = st.progress(0, text="Parsing...")
            for i, uf in enumerate(uploaded_files):
                progress.progress((i+1)/len(uploaded_files), text=f"Parsing {uf.name}...")
                bodies, hw, name = extract_bodies_from_step(uf.read(), uf.name)
                total_hw += hw
                all_groups.update(process_assembly(bodies, name))
            progress.empty()
            st.success(f"Extracted from {len(uploaded_files)} file(s). Filtered {total_hw} hardware components.")

            rows = []
            for gname, gpanels in all_groups.items():
                adj = apply_edgeband_deductions(gpanels, ded_mm, apply_ded)
                for p in adj:
                    cab = re.search(r"(HSW-\d+)", gname)
                    pt = p["part_type"]
                    rows.append({
                        "Section": gname, "Part Description": pt,
                        "Material": material_map.get(pt, mat_side),
                        "Cabinet No.": cab.group(1) if cab else "",
                        "Length (mm)": int(p["length"]), "Width (mm)": int(p["width"]),
                        "Thickness": int(p["thickness"]), "Qty": p["qty"],
                        "Edging Code": p["edging_code"],
                        "Edging Type": default_etype if p["edging_code"] != "0" else "None",
                        "Edging Profile": default_eprof if p["edging_code"] != "0" else "",
                        "Source": "STEP",
                        "Notes": p.get("auto_notes", ""),
                    })
            existing = st.session_state.get("panel_data", []) or []
            st.session_state["panel_data"] = existing + rows
            st.session_state["project_settings"] = project_settings
            st.info(f"Added {len(rows)} STEP rows. Table now has {len(existing) + len(rows)} total.")

    with tab_pdf:
        st.markdown("Upload WCG supplier drawing PDFs (with **Component:** description blocks). "
                    "Each component's dimensions are read from the drawing alongside its block.")
        pdf_files = st.file_uploader(
            "Upload supplier drawing PDFs",
            type=["pdf"], accept_multiple_files=True, key="pdf_upload",
        )
        if pdf_files and st.button("Extract Panels from PDFs", type="primary", key="extract_pdf"):
            all_groups = {}
            total_comps = 0
            progress = st.progress(0, text="Parsing PDFs...")
            for i, uf in enumerate(pdf_files):
                progress.progress((i + 1) / len(pdf_files), text=f"Parsing {uf.name}...")
                try:
                    bodies, name = extract_panels_from_pdf(uf.read(), uf.name)
                except Exception as exc:
                    st.error(f"Failed to parse {uf.name}: {exc}")
                    continue
                total_comps += len(bodies)
                group = process_pdf_assembly(bodies, name)
                for k, v in group.items():
                    if k in all_groups:
                        all_groups[k].extend(v)
                    else:
                        all_groups[k] = v
            progress.empty()
            if total_comps == 0:
                st.warning("No component blocks found. The PDFs may not be in the WCG supplier-sheet format.")
            else:
                st.success(f"Extracted {total_comps} components from {len(pdf_files)} PDF(s).")

            rows = []
            for gname, gpanels in all_groups.items():
                adj = apply_edgeband_deductions(gpanels, ded_mm, apply_ded)
                for p in adj:
                    pt = p["part_type"]
                    pdf_thk = int(p["thickness"]) if p["thickness"] else 18
                    notes = p.get("auto_notes", "")
                    if p.get("component_name"):
                        notes = f"[{p['component_name']}]" + (f"; {notes}" if notes else "")
                    rows.append({
                        "Section": gname, "Part Description": pt,
                        "Material": material_map.get(pt, mat_side),
                        "Cabinet No.": "",
                        "Length (mm)": int(p["length"]),
                        "Width (mm)": int(p["width"]),
                        "Thickness": pdf_thk,
                        "Qty": p["qty"],
                        "Edging Code": p["edging_code"],
                        "Edging Type": default_etype if p["edging_code"] != "0" else "None",
                        "Edging Profile": default_eprof if p["edging_code"] != "0" else "",
                        "Source": "PDF",
                        "Notes": notes,
                    })
            if rows:
                existing = st.session_state.get("panel_data", []) or []
                st.session_state["panel_data"] = existing + rows
                st.session_state["project_settings"] = project_settings
                st.info(f"Added {len(rows)} PDF rows. Table now has {len(existing) + len(rows)} total.")

    with tab_manual:
        st.markdown("Add panels manually for non-standard items.")
        with st.form("manual_add"):
            cols = st.columns([2,2,1,1,1,1,1,1])
            m_part = cols[0].selectbox("Part Type", PART_TYPES, key="mp")
            m_mat = cols[1].selectbox("Material", MATERIAL_LIST, key="mm")
            m_cab = cols[2].text_input("Cabinet", key="mcab", placeholder="HSW-01")
            m_len = cols[3].number_input("Length", min_value=0, step=1, key="ml")
            m_wid = cols[4].number_input("Width", min_value=0, step=1, key="mw")
            m_thk = cols[5].number_input("Thick", min_value=0, value=18, step=1, key="mtk")
            m_qty = cols[6].number_input("Qty", min_value=1, value=1, step=1, key="mq")
            m_edge = cols[7].selectbox("Edging", EDGING_CODES, key="meg")
            if st.form_submit_button("Add Panel") and m_len > 0 and m_wid > 0:
                if "panel_data" not in st.session_state:
                    st.session_state["panel_data"] = []
                st.session_state["panel_data"].append({
                    "Section": "Manual Entry", "Part Description": m_part,
                    "Material": MATERIAL_PRESETS.get(m_mat, m_mat) or m_mat,
                    "Cabinet No.": m_cab, "Length (mm)": int(m_len),
                    "Width (mm)": int(m_wid), "Thickness": int(m_thk),
                    "Qty": int(m_qty), "Edging Code": m_edge,
                    "Edging Type": default_etype if m_edge != "0" else "None",
                    "Edging Profile": default_eprof if m_edge != "0" else "",
                    "Source": "Manual", "Notes": "",
                })
                st.success(f"Added {m_part} ({m_len} x {m_wid} x {m_thk})")

    if "panel_data" in st.session_state and st.session_state["panel_data"]:
        st.divider()
        head_l, head_r = st.columns([4, 2])
        head_l.subheader(f"Panel Schedule - Review & Edit  ({len(st.session_state['panel_data'])} rows)")
        head_l.caption("STEP, PDF and manual rows are combined here. Filter by Source column or use the buttons to clear by origin.")
        with head_r:
            bc1, bc2, bc3, bc4 = st.columns(4)
            if bc1.button("Clear STEP", help="Remove rows added from STEP files"):
                st.session_state["panel_data"] = [r for r in st.session_state["panel_data"] if r.get("Source") != "STEP"]
                st.rerun()
            if bc2.button("Clear PDF", help="Remove rows added from PDF drawings"):
                st.session_state["panel_data"] = [r for r in st.session_state["panel_data"] if r.get("Source") != "PDF"]
                st.rerun()
            if bc3.button("Clear Manual", help="Remove manually-entered rows"):
                st.session_state["panel_data"] = [r for r in st.session_state["panel_data"] if r.get("Source") != "Manual"]
                st.rerun()
            if bc4.button("Clear ALL", type="primary", help="Remove every row"):
                st.session_state["panel_data"] = []
                st.rerun()

        df = pd.DataFrame(st.session_state["panel_data"])
        if "Source" not in df.columns:
            df["Source"] = ""
        column_order = ["Source", "Section", "Part Description", "Material", "Cabinet No.",
                        "Length (mm)", "Width (mm)", "Thickness", "Qty",
                        "Edging Code", "Edging Type", "Edging Profile", "Notes"]
        df = df.reindex(columns=column_order)
        edited_df = st.data_editor(df, column_config={
            "Source": st.column_config.SelectboxColumn("Source", options=["STEP", "PDF", "Manual"], width="small", help="Where this row came from"),
            "Section": st.column_config.TextColumn("Section", width="large"),
            "Part Description": st.column_config.SelectboxColumn("Part Description", options=PART_TYPES, width="medium"),
            "Material": st.column_config.TextColumn("Material", width="large"),
            "Cabinet No.": st.column_config.TextColumn("Cabinet No.", width="small"),
            "Length (mm)": st.column_config.NumberColumn("Length", min_value=0, format="%d"),
            "Width (mm)": st.column_config.NumberColumn("Width", min_value=0, format="%d"),
            "Thickness": st.column_config.NumberColumn("Thick", min_value=0, format="%d"),
            "Qty": st.column_config.NumberColumn("Qty", min_value=1, format="%d"),
            "Edging Code": st.column_config.SelectboxColumn("Edging", options=EDGING_CODES, width="small"),
            "Edging Type": st.column_config.SelectboxColumn("Edging Type", options=EDGING_TYPES, width="medium"),
            "Edging Profile": st.column_config.SelectboxColumn("Edging Profile", options=EDGING_PROFILES, width="small"),
            "Notes": st.column_config.TextColumn("Notes", width="medium"),
        }, num_rows="dynamic", use_container_width=True, key="panel_editor")

        st.session_state["panel_data"] = edited_df.to_dict("records")

        st.divider()
        c1, c2, c3, c4 = st.columns(4)
        total = int(edited_df["Qty"].sum()) if not edited_df.empty else 0
        c1.metric("Total Panels", total)
        c2.metric("Unique Sizes", len(edited_df))
        c3.metric("Sections", edited_df["Section"].nunique() if not edited_df.empty else 0)
        area = (edited_df["Length (mm)"] * edited_df["Width (mm)"] * edited_df["Qty"]).sum() / 1e6 if not edited_df.empty else 0
        c4.metric("Total Area", f"{area:.2f} m2")

        st.divider()
        st.subheader("Export to Smart Space Template")
        if st.button("Generate Smart Space XLSX", type="primary"):
            settings = st.session_state.get("project_settings", project_settings)
            pg = {}
            for _, row in edited_df.iterrows():
                sec = row.get("Section", "Panels")
                pg.setdefault(sec, []).append({
                    "part_type": row.get("Part Description",""), "material": row.get("Material",""),
                    "cabinet_no": row.get("Cabinet No.",""), "length": row.get("Length (mm)",0),
                    "width": row.get("Width (mm)",0), "qty": row.get("Qty",1),
                    "edging_code": row.get("Edging Code","0"), "edging_type": row.get("Edging Type",""),
                    "edging_profile": row.get("Edging Profile",""), "notes": row.get("Notes",""),
                })
            buf = build_smart_space_xlsx(settings, pg)
            fn = f"SMART_SPACE-Panels-{settings.get('project_name','panels').replace(' ','_')}.xlsx"
            st.download_button(f"Download {fn}", buf, fn,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success(f"Ready: {total} panels")
    elif "panel_data" not in st.session_state:
        st.info("Upload STEP files, PDF supplier drawings, or add panels manually to get started.")


if __name__ == "__main__":
    main()
